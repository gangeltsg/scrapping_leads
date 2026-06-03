import uuid
import threading
import json
import os
import csv
import io

from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, Response
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel

from scraper import scrape_domains
from f5bot import monitor_keywords

app = FastAPI()
templates = Jinja2Templates(directory="templates")

# ── Job store (file-based, safe across workers) ────────────────────────────────

_JOBS_DIR = os.path.join(os.environ.get("TMPDIR", "/tmp"), "scraper_jobs")
os.makedirs(_JOBS_DIR, exist_ok=True)


def _job_path(job_id: str) -> str:
    return os.path.join(_JOBS_DIR, f"{job_id}.json")


def _get_job(job_id: str) -> dict | None:
    try:
        with open(_job_path(job_id)) as f:
            return json.load(f)
    except FileNotFoundError:
        return None


def _save_job(job_id: str, data: dict) -> None:
    path = _job_path(job_id)
    tmp = path + ".tmp"
    with open(tmp, "w") as f:
        json.dump(data, f)
    os.replace(tmp, path)


def _update_job(job_id: str, **kwargs) -> None:
    job = _get_job(job_id) or {}
    job.update(kwargs)
    _save_job(job_id, job)


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(request=request, name="index.html")


# ── Scraper ────────────────────────────────────────────────────────────────────

class ScrapeRequest(BaseModel):
    domains: list[str] = []
    keywords: list[str] = []


@app.post("/api/scrape")
async def start_scrape(data: ScrapeRequest):
    raw_domains = [d.strip() for d in data.domains if d.strip()]
    raw_keywords = [k.strip().lower() for k in data.keywords if k.strip()]

    if not raw_domains:
        raise HTTPException(status_code=400, detail="Debes ingresar al menos un dominio.")
    if not raw_keywords:
        raise HTTPException(status_code=400, detail="Debes ingresar al menos un keyword.")

    job_id = str(uuid.uuid4())
    _save_job(job_id, {"status": "running", "results": [], "errors": [], "type": "scrape"})

    threading.Thread(
        target=_run_scrape_job, args=(job_id, raw_domains, raw_keywords), daemon=True
    ).start()

    return JSONResponse(content={"job_id": job_id}, status_code=202)


def _run_scrape_job(job_id: str, domains: list[str], keywords: list[str]) -> None:
    try:
        results, errors = scrape_domains(domains, keywords)
        _update_job(job_id, results=results, errors=errors, status="done")
    except Exception as exc:  # noqa: BLE001
        job = _get_job(job_id) or {}
        job["status"] = "error"
        job.setdefault("errors", []).append(str(exc))
        _save_job(job_id, job)


# ── Monitor (F5Bot / Reddit / HN) ─────────────────────────────────────────────

class MonitorRequest(BaseModel):
    keywords: list[str] = []
    sources: list[str] = ["reddit", "hn"]
    reddit_limit: int = 25
    hn_limit: int = 15


@app.post("/api/monitor")
async def start_monitor(data: MonitorRequest):
    raw_keywords = [k.strip() for k in data.keywords if k.strip()]
    sources = [s for s in data.sources if s in ("reddit", "hn")]

    if not raw_keywords:
        raise HTTPException(status_code=400, detail="Debes ingresar al menos un keyword.")
    if not sources:
        raise HTTPException(status_code=400, detail="Selecciona al menos una fuente.")

    reddit_limit = min(data.reddit_limit, 100)
    hn_limit = min(data.hn_limit, 50)

    job_id = str(uuid.uuid4())
    _save_job(job_id, {"status": "running", "results": [], "errors": [], "type": "monitor"})

    threading.Thread(
        target=_run_monitor_job,
        args=(job_id, raw_keywords, sources, reddit_limit, hn_limit),
        daemon=True,
    ).start()

    return JSONResponse(content={"job_id": job_id}, status_code=202)


def _run_monitor_job(
    job_id: str,
    keywords: list[str],
    sources: list[str],
    reddit_limit: int,
    hn_limit: int,
) -> None:
    try:
        results, errors = monitor_keywords(
            keywords, sources=sources,
            reddit_limit=reddit_limit, hn_limit=hn_limit,
        )
        _update_job(job_id, results=results, errors=errors, status="done")
    except Exception as exc:  # noqa: BLE001
        job = _get_job(job_id) or {}
        job["status"] = "error"
        job.setdefault("errors", []).append(str(exc))
        _save_job(job_id, job)


# ── Status ─────────────────────────────────────────────────────────────────────

@app.get("/api/status/{job_id}")
async def job_status(job_id: str):
    job = _get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado.")
    return job

# ── Export ─────────────────────────────────────────────────────────────────

_SCRAPE_HEADERS = [
    "Dominio", "Empresa", "Descripción", "Dirección", "Tecnologías",
    "Emails", "Teléfonos", "WhatsApp Números",
    "LinkedIn", "Instagram", "Facebook", "Twitter",
    "Keywords Encontrados", "Páginas Escaneadas",
]
_MONITOR_HEADERS = [
    "Fuente", "Keyword", "Título", "URL", "Autor",
    "Subreddit", "Votos", "Comentarios", "Extracto", "Fecha",
]


def _flatten_scrape(r: dict) -> list:
    s = r.get("socials", {})
    return [
        r.get("domain", ""),
        r.get("company_name", ""),
        r.get("description", ""),
        r.get("address", ""),
        " | ".join(r.get("technologies", [])),
        " | ".join(r.get("emails", [])),
        " | ".join(r.get("phones", [])),
        " | ".join(r.get("whatsapp_numbers", [])),
        " | ".join(s.get("linkedin", [])),
        " | ".join(s.get("instagram", [])),
        " | ".join(s.get("facebook", [])),
        " | ".join(s.get("twitter", [])),
        " | ".join(r.get("keywords_found", [])),
        r.get("pages_scraped", 0),
    ]


def _flatten_monitor(r: dict) -> list:
    return [
        r.get("source", ""),
        r.get("keyword", ""),
        r.get("title", ""),
        r.get("url", ""),
        r.get("author", ""),
        r.get("subreddit", ""),
        r.get("score", 0),
        r.get("num_comments", 0),
        r.get("selftext_snippet", ""),
        r.get("created", ""),
    ]


@app.get("/api/export/{job_id}")
async def export_job(job_id: str, format: str = "csv"):
    job = _get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado.")
    if job["status"] != "done":
        raise HTTPException(status_code=400, detail="El job aún no ha terminado.")

    results = job.get("results", [])
    job_type = job.get("type", "scrape")
    headers = _SCRAPE_HEADERS if job_type == "scrape" else _MONITOR_HEADERS
    flatten = _flatten_scrape if job_type == "scrape" else _flatten_monitor
    rows = [flatten(r) for r in results]
    fname = f"leads_{job_type}"

    if format == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads Web" if job_type == "scrape" else "Menciones"
        ws.append(headers)
        for row in rows:
            ws.append(row)
        hfont = Font(bold=True, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="4F46E5")
        for cell in ws[1]:
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 60)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'},
        )

    # CSV (utf-8-sig so Excel opens it correctly)
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(headers)
    writer.writerows(rows)
    return Response(
        content=out.getvalue().encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{fname}.csv"'},
    )

# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8001))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=False)
